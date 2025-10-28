from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort, jsonify
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import os, time, base64, re

from db import (
    sp_equipo_upsert,
    sp_equipo_agregar_cambio,
    sp_historial_por_persona,
    query_dispositivos,
    historial_por_equipo,
    sp_equipo_reasignar,
    sp_equipo_dar_baja,
    equipo_upsert_completo,
    obtener_equipo_por_tag,
    archivo_principal_get,
    archivo_principal_set,
)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'cambia_esto_mel'
app.url_map.strict_slashes = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.config['ENABLE_UPLOAD'] = False

SHARE_ROOT = r'\\itzamna\DATAUSERS\ADM Y SISTEMAS\Entrega Equipos'
UPLOAD_ROOT = r'\\itzamna\DATAUSERS\ADM Y SISTEMAS\Entrega Equipos'

ALLOWED_EXTS = {'pdf', 'jpg', 'jpeg', 'png', 'xlsx', 'xls'}


def _allowed(name: str) -> bool:
    return '.' in name and name.rsplit('.', 1)[-1].lower() in ALLOWED_EXTS


def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def _is_inside(path: str, base: str) -> bool:
    try:
        return os.path.commonpath([os.path.abspath(path), os.path.abspath(base)]) == os.path.abspath(base)
    except Exception:
        return False


def _encode_path(p: str) -> str:
    return base64.urlsafe_b64encode(p.encode('utf-8')).decode('ascii')


def _decode_path(tok: str) -> str:
    return base64.urlsafe_b64decode(tok.encode('ascii')).decode('utf-8')


def _scan_files_for_tag(tag: str):
    out = []
    tag_l = (tag or '').lower()
    if not tag_l or not os.path.isdir(SHARE_ROOT):
        return out
    for root, _, files in os.walk(SHARE_ROOT):
        for nm in files:
            if tag_l in nm.lower() and _allowed(nm):
                full = os.path.join(root, nm)
                try:
                    size = os.path.getsize(full)
                    mtime = time.strftime('%Y-%m-%d %H:%M', time.localtime(os.path.getmtime(full)))
                except OSError:
                    continue
                out.append({"name": nm, "path": full, "size": size, "mtime": mtime, "token": _encode_path(full)})
    out.sort(key=lambda x: (x["mtime"], x["name"]), reverse=True)
    return out


def save_equipo_file_principal(tag: str, file_storage):
    if not file_storage or not file_storage.filename:
        return None, "Archivo vacío."
    fname = secure_filename(file_storage.filename)
    if not _allowed(fname):
        return None, "Tipo de archivo no permitido."

    _ensure_dir(UPLOAD_ROOT)
    ext = os.path.splitext(fname)[1].lower()
    stored = f"{tag}{ext}"
    path = os.path.join(UPLOAD_ROOT, stored)

    try:
        file_storage.stream.seek(0, os.SEEK_END)
        size = file_storage.stream.tell()
        file_storage.stream.seek(0)
    except Exception:
        size = None
    if size is not None and size > 5 * 1024 * 1024:
        return None, "Máximo 5MB por archivo."

    file_storage.save(path)
    archivo_principal_set(tag, path, stored)
    return stored, None


# ---------- Helpers para autollenado desde Excel ----------
def _xlsx_to_grid(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    grid = []
    for r in ws.iter_rows(values_only=True):
        row = []
        for v in r:
            s = '' if v is None else str(v).strip()
            row.append(s)
        grid.append(row)
    return grid


def _find_cell(grid, needle):
    n = (needle or '').lower()
    for r, row in enumerate(grid):
        for c, val in enumerate(row):
            if n in (val or '').lower():
                return (r, c)
    return None


def _safe_get(grid, r, c):
    try:
        return grid[r][c]
    except Exception:
        return ''


def _truthy(s):
    s = (s or '').strip().lower()
    return s in {'✔', 'si', 'sí', 'true', '1', 'x'}


def _extract_tag_from_text(text):
    m = re.search(r'(?i)\bactivo\b\s*[:#-]?\s*([A-Za-z0-9\-]+)', text or '')
    return m.group(1).strip() if m else ''


# =========================
# RUTAS
# =========================
@app.route('/')
def index():
    q = request.args.get('q', '').strip()
    person = request.args.get('person', '').strip()
    dispositivos = query_dispositivos(
        filtro_tag=q or None,
        filtro_persona=person or None,
        solo_activos=True
    )
    return render_template('index.html', dispositivos=dispositivos, q=q, person=person)


# --- Nuevo equipo (con vínculo/subida del archivo principal opcional) ---
@app.route('/device/new', methods=['GET', 'POST'])
@app.route('/new', methods=['GET', 'POST'])
def new_device():
    if request.method == 'POST':
        tag = request.form['tag']
        marca = request.form.get('marca')
        modelo = request.form.get('modelo')

        # ✅ Tipo de equipo (Portátil, Todo en uno, etc.)
        tipo_equipo = request.form.get('tipo_equipo') or None

        serial = request.form.get('serial')
        ubicacion = request.form.get('ubicacion')
        persona_asignada = request.form.get('persona_asignada')

        # ✅ Tipo de asignación
        tipo_ubicacion = request.form.get('tipo_ubicacion') or 'ALMACEN'
        area = request.form.get('area') or None
        cargo = request.form.get('cargo') or None

        cargador = 1 if request.form.get('cargador') else 0
        maletin = 1 if request.form.get('maletin') else 0
        mouse = 1 if request.form.get('mouse') else 0
        teclado = 1 if request.form.get('teclado') else 0
        impresora = 1 if request.form.get('impresora') else 0
        lector = 1 if request.form.get('lector') else 0
        observaciones = request.form.get('observaciones')

        # ✅ Pasar tipo_equipo correctamente
        equipo_upsert_completo(
            tag, marca, modelo, serial, ubicacion, persona_asignada,
            cargador, maletin, mouse, teclado, observaciones,
            impresora, lector, tipo_equipo, tipo_ubicacion, area, cargo
        )

        # --- ARCHIVO PRINCIPAL (sin cambios) ---
        ruta_manual = (request.form.get('ruta_principal') or '').strip()
        archivo_subido = request.files.get('archivo') if app.config.get('ENABLE_UPLOAD') else None

        if ruta_manual:
            if not _is_inside(ruta_manual, SHARE_ROOT):
                flash('La ruta pegada no pertenece a la carpeta de red configurada.', 'danger')
            elif not os.path.isfile(ruta_manual):
                flash('La ruta pegada no existe como archivo.', 'danger')
            elif not _allowed(os.path.basename(ruta_manual)):
                flash('Extensión de archivo no permitida.', 'danger')
            else:
                archivo_principal_set(tag, ruta_manual, os.path.basename(ruta_manual))
                flash('Archivo principal vinculado correctamente.', 'success')
        elif archivo_subido and archivo_subido.filename:
            stored, err = save_equipo_file_principal(tag, archivo_subido)
            if err:
                flash(f'Archivo no subido: {err}', 'danger')
            else:
                flash('Archivo principal subido y vinculado.', 'success')

        flash('Equipo registrado/actualizado', 'success')
        return redirect(url_for('index'))

    return render_template('new_device.html', enable_upload=app.config['ENABLE_UPLOAD'])


# === AUTOFILL desde ruta Excel (POST JSON) ===
@app.post('/new/autofill_from_path')
def new_autofill_from_path():
    ruta = (request.json or {}).get('ruta', '').strip()
    if not ruta:
        return {"ok": False, "error": "No enviaste la ruta."}, 400
    if not _is_inside(ruta, SHARE_ROOT):
        return {"ok": False, "error": "La ruta no pertenece a la carpeta de red configurada."}, 400
    if not os.path.isfile(ruta):
        return {"ok": False, "error": "La ruta no existe como archivo."}, 404
    if not ruta.lower().endswith(('.xlsx', '.xls')):
        return {"ok": False, "error": "Solo se admite Excel (.xlsx/.xls) para autollenar."}, 400

    try:
        grid = _xlsx_to_grid(ruta)
    except Exception as e:
        return {"ok": False, "error": f"No pude leer el Excel: {e}"}, 400

    # debug en servidor para ver el contenido (opcional)
    print("\n=== CONTENIDO DEL EXCEL (primeras 30 filas) ===")
    for i, row in enumerate(grid[:30]):
        print(f"Fila {i}: {row}")
    print("=" * 50 + "\n")

    def find_cell(needle):
        pos = _find_cell(grid, needle)
        print(f"[DEBUG] Buscando '{needle}': {pos}")
        return pos

    def get(r, c):
        return _safe_get(grid, r, c)

    def col_of(header, header_row):
        h = header.lower()
        for c, val in enumerate(grid[header_row]):
            if h in (val or '').lower():
                return c
        return None

    def first_below(col, header_row, max_down=6):
        for r in range(header_row + 1, min(header_row + 1 + max_down, len(grid))):
            txt = get(r, col)
            if txt:
                return txt
        return ''

    def first_right(row, start_c, max_right=10):
        if row < 0 or row >= len(grid):
            return ''
        for cc in range(start_c + 1, min(start_c + 1 + max_right, len(grid[row]))):
            txt = get(row, cc)
            if txt:
                return txt
        return ''

    data = {
        "tipo_equipo": "",
        "marca": "",
        "modelo": "",
        "serial": "",
        "ubicacion": "",
        "persona_asignada": "",
        "observaciones": "",
        "cargador": 0,
        "maletin": 0,
        "mouse": 0,
        "teclado": 0,
        "tag": "",
        "area": "",
        "cargo": ""
    }

    # Buscar encabezado "Equipo Entregado" para columnas
    pos_eq = find_cell('Equipo Entregado')
    if pos_eq:
        r_head, _ = pos_eq
        c_tipo = col_of('Equipo Entregado', r_head)
        c_marca = col_of('Marca', r_head)
        c_modelo = col_of('Modelo', r_head)
        c_serial = col_of('Serial', r_head)

        if c_tipo is not None:
            data["tipo_equipo"] = first_below(c_tipo, r_head)
        if c_marca is not None:
            data["marca"] = first_below(c_marca, r_head)
        if c_modelo is not None:
            data["modelo"] = first_below(c_modelo, r_head)
        if c_serial is not None:
            data["serial"] = first_below(c_serial, r_head)

    # Ubicación (derecha de "Lugar")
    pos_lugar = find_cell('Lugar')
    if pos_lugar:
        r, c = pos_lugar
        data["ubicacion"] = get(r, c + 1)

    # Persona asignada (buscar 'Entregado a' y luego 'Nombre' cercano)
    pos_entregado_a = find_cell('Entregado a')
    if pos_entregado_a:
        rA, cA = pos_entregado_a
        nombre_rc = None
        for rr in range(rA, min(rA + 12, len(grid))):
            for cc, val in enumerate(grid[rr]):
                if 'nombre' in (val or '').lower():
                    nombre_rc = (rr, cc)
                    break
            if nombre_rc:
                break
        if nombre_rc:
            rn, cn = nombre_rc
            nombre_val = get(rn, cn + 1)
            print(f"[DEBUG] Nombre encontrado en ({rn},{cn}): '{nombre_val}'")
            data["persona_asignada"] = nombre_val

    # si no lo encontramos con la búsqueda anterior, intentar 'Nombre' suelto
    if not data["persona_asignada"]:
        pos_nombre = find_cell('Nombre:') or find_cell('Nombre')
        if pos_nombre:
            r, c = pos_nombre
            nombre_val = first_right(r, c, max_right=8)
            if not nombre_val:
                nombre_val = first_below(c, r, max_down=3)
            print(f"[DEBUG] Nombre (directo) en ({r},{c}): '{nombre_val}'")
            data["persona_asignada"] = nombre_val

    # Área
    pos_area = find_cell('Área:') or find_cell('Área') or find_cell('Area:') or find_cell('Area')
    if pos_area:
        r, c = pos_area
        area_val = first_right(r, c, max_right=8)
        if not area_val:
            area_val = first_below(c, r, max_down=3)
        print(f"[DEBUG] Área encontrada en ({r},{c}): '{area_val}'")
        data["area"] = area_val
    else:
        print("[DEBUG] No se encontró 'Área' en el documento")

    # Cargo
    pos_cargo = find_cell('Cargo:') or find_cell('Cargo')
    if pos_cargo:
        r, c = pos_cargo
        cargo_val = first_right(r, c, max_right=8)
        if not cargo_val:
            cargo_val = first_below(c, r, max_down=3)
        print(f"[DEBUG] Cargo encontrado en ({r},{c}): '{cargo_val}'")
        data["cargo"] = cargo_val
    else:
        print("[DEBUG] No se encontró 'Cargo' en el documento")

    # Observaciones y extracción de tag
    pos_obs = find_cell('Observaciones')
    if pos_obs:
        r, c = pos_obs
        obs = first_right(r, c, max_right=12)
        if not obs:
            obs = first_below(c, r, max_down=6)
        data["observaciones"] = obs
        tag_from_obs = _extract_tag_from_text(obs)
        if tag_from_obs and not data.get("tag"):
            data["tag"] = tag_from_obs

    # Accesorios
    pos_cargador = find_cell('Cargador')
    pos_maletin = find_cell('Maletín') or find_cell('Maletin')
    pos_mouse = find_cell('Mouse')
    pos_teclado = find_cell('Teclado')

    if pos_cargador:
        r, c = pos_cargador
        data["cargador"] = 1 if _truthy(get(r + 1, c)) else 0
    if pos_maletin:
        r, c = pos_maletin
        data["maletin"] = 1 if _truthy(get(r + 1, c)) else 0
    if pos_mouse:
        data["mouse"] = 1
    if pos_teclado:
        data["teclado"] = 1

    print(f"[DEBUG] Datos finales: {data}")
    return {"ok": True, "data": data}


# --- Vista dispositivo ---
@app.route('/device/<tag>', methods=['GET', 'POST'])
def device_view(tag):
    if request.method == 'POST':
        tipo = request.form['tipo']
        desc = request.form.get('descripcion') or None
        fecha = request.form.get('fecha') or None
        persona_rel = request.form.get('persona_rel') or None
        registrado_por = request.form.get('registrado_por') or ''
        sp_equipo_agregar_cambio(tag, tipo, desc, fecha, persona_rel, registrado_por)
        flash('Cambio registrado', 'success')
        return redirect(url_for('device_view', tag=tag))

    historial = historial_por_equipo(tag)
    equipo = obtener_equipo_por_tag(tag)
    principal = archivo_principal_get(tag)

    if principal:
        files_preview = []
    else:
        files_preview = _scan_files_for_tag(tag)[:5]

    return render_template(
        'device.html',
        tag=tag,
        historial=historial,
        equipo=equipo,
        principal=principal,
        files_preview=files_preview,
        enable_upload=app.config['ENABLE_UPLOAD']
    )


# --- Vincular archivo existente de la red ---
@app.get('/device/<tag>/link')
def device_link(tag):
    candidates = _scan_files_for_tag(tag)
    principal = archivo_principal_get(tag)
    return render_template('device_link.html', tag=tag, candidates=candidates, principal=principal)


@app.post('/device/<tag>/link')
def device_link_save(tag):
    token = request.form.get('token') or ''
    ruta_manual = (request.form.get('ruta_manual') or '').strip()

    if token:
        full = _decode_path(token)
    else:
        full = ruta_manual

    if not full:
        flash('Debes seleccionar o pegar una ruta.', 'danger')
        return redirect(url_for('device_link', tag=tag))

    if not _is_inside(full, SHARE_ROOT):
        flash('La ruta no pertenece a la carpeta de red configurada.', 'danger')
        return redirect(url_for('device_link', tag=tag))

    if not os.path.isfile(full):
        flash('La ruta no existe como archivo.', 'danger')
        return redirect(url_for('device_link', tag=tag))

    nombre = os.path.basename(full)
    if not _allowed(nombre):
        flash('Extensión de archivo no permitida.', 'danger')
        return redirect(url_for('device_link', tag=tag))

    archivo_principal_set(tag, full, nombre)
    flash('Archivo principal vinculado correctamente.', 'success')
    return redirect(url_for('device_view', tag=tag))


@app.get('/device/<tag>/principal/download')
def device_principal_download(tag):
    principal = archivo_principal_get(tag)
    if not principal:
        abort(404)
    path = principal.get('ruta')
    name = principal.get('nombre') or os.path.basename(path)
    if not (path and os.path.isfile(path)):
        abort(404)
    return send_file(path, as_attachment=True, download_name=name)


@app.post('/device/<tag>/upload')
def device_upload(tag):
    f = request.files.get('archivo')
    if not f or not f.filename:
        flash('Debes seleccionar un archivo.', 'danger')
        return redirect(url_for('device_view', tag=tag))

    stored, err = save_equipo_file_principal(tag, f)
    if err:
        flash(err, 'danger')
    else:
        flash('Archivo principal subido y vinculado.', 'success')
    return redirect(url_for('device_view', tag=tag))


@app.post('/device/<tag>/reasignar')
def device_reassign(tag):
    nueva = (request.form.get('nueva_persona') or '').strip()
    cargo = request.form.get('cargo') or None
    fecha = request.form.get('fecha') or None
    registrado_por = request.form.get('registrado_por') or ''
    desc = request.form.get('descripcion') or None
    if not nueva:
        flash('Debe indicar el nombre de la nueva persona.', 'danger')
        return redirect(url_for('device_view', tag=tag))
    sp_equipo_reasignar(tag, nueva, cargo, fecha, registrado_por, desc)
    flash(f'Equipo {tag} reasignado a {nueva}.', 'success')
    return redirect(url_for('device_view', tag=tag))


@app.post('/device/<tag>/baja')
def device_baja(tag):
    motivo = request.form.get('motivo') or None
    fecha = request.form.get('fecha') or None
    registrado_por = request.form.get('registrado_por') or ''
    sp_equipo_dar_baja(tag, motivo, fecha, registrado_por)
    flash(f'Equipo {tag} marcado en BAJA.', 'success')
    return redirect(url_for('device_view', tag=tag))


@app.get('/device/<tag>/files')
def device_files(tag):
    principal = archivo_principal_get(tag)
    files = []
    if principal:
        p = principal['ruta']
        if os.path.isfile(p):
            files.append({
                'name': principal['nombre'],
                'size': os.path.getsize(p),
                'mtime': time.strftime('%Y-%m-%d %H:%M', time.localtime(os.path.getmtime(p))),
                'where': 'PRINCIPAL'
            })
    return render_template('device_files.html', tag=tag, files=files)


@app.get('/device/<tag>/files/PRINCIPAL/<path:fname>')
def device_download(tag, fname):
    principal = archivo_principal_get(tag)
    if not principal or principal['nombre'] != fname:
        abort(404)
    path = principal['ruta']
    if not os.path.isfile(path):
        abort(404)
    return send_file(path, as_attachment=True, download_name=fname)


# --- Editar equipo ---
@app.route('/device/<tag>/edit', methods=['GET', 'POST'])
def edit_device(tag):
    equipo = obtener_equipo_por_tag(tag)
    if not equipo:
        flash('Equipo no encontrado.', 'danger')
        return redirect(url_for('device_view', tag=tag))

    if request.method == 'POST':
        marca = request.form.get('marca')
        modelo = request.form.get('modelo')
        serial = request.form.get('serial')
        ubicacion = request.form.get('ubicacion')
        persona_asignada = request.form.get('persona_asignada')

        # ✅ Nuevo: leer tipo_equipo correctamente (Portátil, Todo en uno, etc.)
        tipo_equipo = request.form.get('tipo_equipo') or None

        # ✅ Tipo de asignación (ALMACÉN / ADMINISTRATIVO)
        tipo_ubicacion = request.form.get('tipo_ubicacion') or 'ALMACEN'
        area = request.form.get('area') or None
        cargo = request.form.get('cargo') or None

        cargador = 1 if request.form.get('cargador') else 0
        maletin = 1 if request.form.get('maletin') else 0
        mouse = 1 if request.form.get('mouse') else 0
        teclado = 1 if request.form.get('teclado') else 0
        impresora = 1 if request.form.get('impresora') else 0
        lector = 1 if request.form.get('lector') else 0
        observaciones = request.form.get('observaciones')

        # ✅ Pasar tipo_equipo en el orden correcto
        equipo_upsert_completo(
            tag, marca, modelo, serial, ubicacion, persona_asignada,
            cargador, maletin, mouse, teclado, observaciones,
            impresora, lector, tipo_equipo, tipo_ubicacion, area, cargo
        )

        flash('Equipo actualizado correctamente.', 'success')
        return redirect(url_for('device_view', tag=tag))

    return render_template('edit_device.html', equipo=equipo, tag=tag)


@app.route('/search/person', methods=['GET'])
def search_person():
    nombre = request.args.get('nombre', '').strip()
    resultados = sp_historial_por_persona(nombre) if nombre else []
    return render_template('index.html', resultados_persona=resultados, q='', person=nombre, dispositivos=[])


@app.route('/bajas')
def bajas():
    dispositivos = query_dispositivos(solo_activos=False)
    dispositivos_baja = [d for d in dispositivos if (d.get('estado') or '').upper() == 'BAJA']
    return render_template('bajas.html', dispositivos=dispositivos_baja)


@app.route('/_routes')
def _routes():
    return '<pre>' + '\n'.join(str(r) for r in app.url_map.iter_rules()) + '</pre>'


@app.errorhandler(404)
def err404(e):
    print(f"[404] PATH solicitado: {repr(request.path)}")
    return "Not Found", 404


if __name__ == '__main__':
    print('CWD:', os.getcwd())
    print('URL MAP:\n', app.url_map)
    app.run(debug=True)
